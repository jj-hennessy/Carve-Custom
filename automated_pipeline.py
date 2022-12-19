"""
DESCRIPTION:
------------
This script reads a survey data CSV from Qualtrics and produces a unique recommendation XLSX file
for each row in that CSV. The survey data is plugged into the driver row of a template Excel
recommendation file (with formulas installed), re-runs formulas on provided survey data (file must be re-opened to trigger re-run),
then saves the file in the following format:
"{FIRST NAME}'s Recommendation - {YEAR OF SKI} {NAME OF TOP RECOMMENDED SKI} {SIZE OF SKI}".xlsx
EX: "Jonathan's Recommendation - 2022  Armada ARV 106 180 cm"
"""
import time
import os
import stat
import pandas as pd
import xlwings as xw
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from create_email import send_email
import requests
import shutil

def return_store_link(rec_sheet, cell):
    link = 'https://www.google.com'
    try:
        link = rec_sheet.range(cell).hyperlink
    except:
        print("hyperlink not found")
    return link


start_time = time.time()

# Global variables
DRIVER_SURVEY_ROW = 5

# Read user survey CSV from Qualtrics
survey_df = pd.read_csv('CarveApril6Exp.csv') # NOTE: static file at this point, will change later
survey_df.fillna('', inplace=True) # clean data
"""
ORDERED COLUMN NAMES (44 cols):
----------------------------------------------------------------
'StartDate', 'EndDate', 'Status', 'IPAddress', 'Progress',
'Duration (in seconds)', 'Finished', 'RecordedDate', 'ResponseId',
'RecipientLastName', 'RecipientFirstName', 'RecipientEmail',
'ExternalReference', 'LocationLatitude', 'LocationLongitude',
'DistributionChannel', 'UserLanguage', 'Q36_4', 'Q36_5', 'Q2', 'Q25',
'Q3_1', 'Q3_2', 'Q3_3', 'Q3_4', 'Q3_5', 'Q4_NPS_GROUP', 'Q4', 'Q14_1',
'Q13_1', 'Q13_2', 'Q13_3', 'Q13_4', 'Q13_5', 'Q26', 'Q16', 'Q23',
'Q23_3_TEXT', 'Q28', 'Q17', 'Q29', 'Q35_7', 'Q35_42', 'Q35_43'
----------------------------------------------------------------
"""

# EXAMPLE DICTIONARIES:
# {'StartDate': '2/28/22 0:34', 'EndDate': '2/28/22 0:37', 'Status': 'IP Address', 'IPAddress': '76.224.184.181', 'Progress': 100, 'Duration (in seconds)': 172, 'Finished': True, 'RecordedDate': '2/28/22 0:37', 'ResponseId': 'R_3lQOXbNIRcHA6cx', 'RecipientLastName': '', 'RecipientFirstName': '', 'RecipientEmail': '', 'ExternalReference': '', 'LocationLatitude': 37.92489624, 'LocationLongitude': -122.5099945, 'DistributionChannel': 'anonymous', 'UserLanguage': 'EN', 'Q36_4': "6' 4''", 'Q36_5': 205.0, 'Q2': 'Ski', 'Q25': "Unisex (sometimes called Men's skis)", 'Q3_1': 51, 'Q3_2': '', 'Q3_3': 10.0, 'Q3_4': 40, 'Q3_5': '', 'Q4_NPS_GROUP': 'Detractor', 'Q4': 3.0, 'Q14_1': 'Palisades Tahoe', 'Q13_1': 20.0, 'Q13_2': 50.0, 'Q13_3': 20.0, 'Q13_4': 10.0, 'Q13_5': '', 'Q26': 'No', 'Q16': '', 'Q23': 'Show me everything (default)', 'Q23_3_TEXT': '', 'Q28': 'No, I mostly care about performance', 'Q17': '', 'Q29': '', 'Q35_7': 'Peter ', 'Q35_42': 'Papi@squeeril.com', 'Q35_43': '415-924-7894'}
# {'StartDate': '2/27/22 23:37', 'EndDate': '2/27/22 23:40', 'Status': 'IP Address', 'IPAddress': '209.52.88.216', 'Progress': 100, 'Duration (in seconds)': 207, 'Finished': True, 'RecordedDate': '2/27/22 23:40', 'ResponseId': 'R_3kBUS4zDbu6hWUt', 'RecipientLastName': '', 'RecipientFirstName': '', 'RecipientEmail': '', 'ExternalReference': '', 'LocationLatitude': 49.3249054, 'LocationLongitude': -122.8628006, 'DistributionChannel': 'anonymous', 'UserLanguage': 'EN', 'Q36_4': "5' 11''", 'Q36_5': 160.0, 'Q2': 'Ski', 'Q25': "Unisex (sometimes called Men's skis)", 'Q3_1': 38, 'Q3_2': 15.0, 'Q3_3': 50.0, 'Q3_4': 46, 'Q3_5': 35.0, 'Q4_NPS_GROUP': 'Detractor', 'Q4': 5.0, 'Q14_1': 'Grouse Mountain', 'Q13_1': 14.0, 'Q13_2': 59.0, 'Q13_3': 36.0, 'Q13_4': 13.0, 'Q13_5': 37.0, 'Q26': 'Yes', 'Q16': '', 'Q23': 'Show me everything (default)', 'Q23_3_TEXT': '', 'Q28': 'Yes', 'Q17': 'Red,Orange,Yellow,Green,Blue,Purple,Pink,White,Black,Brown', 'Q29': '\nAbstract,\nImagery,\nPatterned,\nFlat color', 'Q35_7': 'Gavin', 'Q35_42': 'Gavgt2@gmail.com', 'Q35_43': '6043129575'}

# Read in template workbook
tmp_workbook = load_workbook(filename="rec_template.xlsx")
tmp_driver_sheet = tmp_workbook['Driver']

cwd_save_dir = 'carve_recs'
recs_path = f"{os.getcwd()}/{cwd_save_dir}"
Path(recs_path).mkdir(parents=True, exist_ok=True)

# Generate new recommendation XLSX file for each row in survey CSV
survey_df_dict = survey_df.to_dict('records') # fastest way to iterate through dataframe
image_links = ['B48','E48','H48','K48','N48','Q48']
for row in survey_df_dict:
    # Insert survey data from column A (0) to AR (43)
    for col_index, col_name in enumerate(row.keys()):
        tmp_driver_sheet.cell(column=col_index+1, row=DRIVER_SURVEY_ROW, value=row[col_name])

    # Create save path + file name & make save dir if doesn't exist
    first_name = row['Q35_7'].lower().capitalize().strip(" ")
    email = row['Q35_42']
    file_name = f"{first_name}'s Recommendation.xlsx"
    rec_file_path = f"{recs_path}/{file_name}"
    tmp_workbook.save(rec_file_path)
    tmp_workbook.close()
    os.chmod(rec_file_path, stat.S_IRWXU|stat.S_IRWXG|stat.S_IRWXO)

    # Grab this user's top recommended ski
    rec_book  = xw.Book(rec_file_path)
    rec_sheet = rec_book.sheets['Your recommendation']
    num_skis = rec_sheet.range('C18').value.split(" ")[0]
    top_ski = rec_sheet.range('B35').value
    gender = rec_sheet.range('B36').value
    image = rec_sheet.range('B48').value

    # Get the images of the skis
    # flagContinue = False
    # for cell_name in image_links:
    #     file_name = cell_name + ".jpg"
    #     image_link = rec_sheet.range(cell_name).value
    #     rng = rec_sheet.range(cell_name)
    #     image_link = rng.value
    #     rng.value = ''
    #     if image_link == 0.0:
    #         continue
    #     try:
    #         res = requests.get(image_link, stream = True)
    #     except:
    #         flagContinue = True
    #         break
    #     if res.status_code == 200:
    #         with open(file_name,'wb') as f:
    #             shutil.copyfileobj(res.raw, f)
    #         rec_sheet.pictures.add(file_name, top=rng.top, left=rng.left, width = 280, height = 250)
    #         print('Image sucessfully Downloaded: ',file_name)
    #     else:
    #         print('Image Couldn\'t be retrieved')
    # if flagContinue:
    #     record_file = open('record_responses.csv', 'a')
    #     record_file.write(f'{first_name},{email},,,,\n')
    #     record_file.close()
    #     continue
    # how to extract hyperlink
    # print(rec_sheet.range('B38').hyperlink)

    sellers = []
    if len(str(rec_sheet.range('B38').value)) > 1 and '0.0' not in str(rec_sheet.range('B38').value):
        sellers.append([rec_sheet.range('B38').value, rec_sheet.range('C38').value, rec_sheet.range('BD75').value])

    if len(str(rec_sheet.range('B39').value)) > 1 and '0.0' not in str(rec_sheet.range('B39').value):
        sellers.append([rec_sheet.range('B39').value, rec_sheet.range('C39').value, rec_sheet.range('BH75').value])

    if len(str(rec_sheet.range('B40').value)) > 1 and '0.0' not in str(rec_sheet.range('B40').value):
        sellers.append([rec_sheet.range('B40').value, rec_sheet.range('C40').value, rec_sheet.range('BL75').value])

    overall = rec_sheet.range('C42').value
    speed = rec_sheet.range('C43').value
    man = rec_sheet.range('C44').value
    rec_book.save()
    rec_book.close()

    try:
        name, receiver_email, seller_info_sent, picture_sent, sent_email_a = send_email(first_name, num_skis, top_ski, image, gender, sellers, float(overall) * 100, float(speed) * 100, float(man) * 100, email)
        record_file = open('record_responses.csv', 'a')
        record_file.write(f'{name},{receiver_email},{seller_info_sent},{picture_sent},{sent_email_a},\n')
        record_file.close()
    except Exception as e:
        print(f'Error sending email for {name} is {str(e)}')
        record_file = open('record_responses.csv', 'a')
        record_file.write(f'{first_name},{email},,,,{str(e)}\n')
        record_file.close()



# Analytics
total_runtime = time.time() - start_time
print(f"\nTotal Runtime: {int(total_runtime/60)} minutes {int(total_runtime%60)} seconds")
print(f"Users Processed: {len(survey_df.index)}")
print(f"Avg. Processing Time Per User: {int(total_runtime/len(survey_df.index))} seconds")
